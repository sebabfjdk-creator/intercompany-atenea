"""Endpoints del módulo Conciliación Bancaria."""
from __future__ import annotations

import io
import os
import tempfile

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.services import bank_service as svc
from db.base import get_db
from db.models import AuditLog, BankAccount, User

router = APIRouter(prefix="/api/bancos", tags=["bancos"])


def _puede_escribir(db: Session, user: User) -> None:
    acc = db.scalars(select(BankAccount)).first()
    sistema = acc.sistema if acc else "ES"
    if sistema == "ES" and user.rol == "admin_co":
        raise HTTPException(403, "admin_co no puede modificar conciliaciones de cuentas de España")


async def _save_tmp(file: UploadFile) -> str:
    suffix = os.path.splitext(file.filename or "")[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    tmp.write(await file.read())
    tmp.close()
    return tmp.name


async def _ingest(origen: str, file: UploadFile, db: Session, user: User) -> dict:
    _puede_escribir(db, user)
    path = await _save_tmp(file)
    try:
        try:
            res = svc.ingest(db, origen, path, user.id)
        except Exception as e:  # noqa: BLE001
            raise HTTPException(422, f"Error procesando el archivo: {e}")
        db.add(AuditLog(entidad="banco", entidad_id=f"{origen}:{res['mes']}", accion="create",
                        valor_despues=str(res), usuario_id=user.id))
        db.commit()
        return {"ok": True, "archivo": file.filename, **res}
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass


@router.post("/contable")
async def subir_contable(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return await _ingest("contable", file, db, user)


@router.post("/extracto")
async def subir_extracto(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return await _ingest("extracto", file, db, user)


@router.get("/periodos")
def periodos(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return svc.periodos(db)


@router.get("/conciliacion")
def conciliacion(mes: str | None = Query(None), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return svc.conciliacion(db, mes)


class SaldosIn(BaseModel):
    mes: str
    saldo_contable: float = 0
    saldo_banco: float = 0


@router.put("/saldos")
def saldos(body: SaldosIn, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _puede_escribir(db, user)
    return svc.set_saldos(db, body.mes, body.saldo_contable, body.saldo_banco)


@router.post("/cerrar")
def cerrar(mes: str = Query(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _puede_escribir(db, user)
    res = svc.cerrar(db, mes, user.id)
    db.add(AuditLog(entidad="banco", entidad_id=mes, accion="close", usuario_id=user.id))
    db.commit()
    return res


@router.post("/reabrir")
def reabrir(mes: str = Query(...), db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    _puede_escribir(db, user)
    res = svc.reabrir(db, mes, user.id)
    db.add(AuditLog(entidad="banco", entidad_id=mes, accion="update", valor_despues="reabierta", usuario_id=user.id))
    db.commit()
    return res


def _money(ws, cell):
    ws[cell].number_format = "#,##0.00;[Red]-#,##0.00"


@router.get("/export")
def export(mes: str | None = Query(None), db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    d = svc.conciliacion(db, mes)
    if d.get("vacio"):
        raise HTTPException(404, "No hay conciliación para exportar")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    bc, bb, bk = d["bloque_contable"], d["bloque_banco"], d["bloque_conciliar"]
    ws.append([f"Conciliación bancaria {d['cuenta']} — {d['mes']} ({d['estado']})"])
    ws.append([])
    ws.append(["SALDOS CONTABLES (LIBROS)"])
    ws.append(["Saldo inicial", bc["inicial"]]); ws.append(["(+) Débitos / ingresos", bc["debito"]])
    ws.append(["(−) Créditos / pagos", bc["credito"]]); ws.append(["Saldo final contable", bc["final"]])
    ws.append([])
    ws.append(["SALDOS BANCARIOS (EXTRACTO)"])
    ws.append(["Saldo inicial", bb["inicial"]]); ws.append(["(+) Ingresos", bb["ingresos"]])
    ws.append(["(−) Egresos", bb["egresos"]]); ws.append(["Saldo final bancario", bb["final"]])
    ws.append([])
    ws.append(["SALDO POR CONCILIAR"])
    ws.append(["Saldo en libros", bk["saldo_libros"]])
    ws.append(["(+) Ingresos extracto no en libros", bk["ing_no_libros"]])
    ws.append(["(−) Egresos extracto no en libros", bk["egr_no_libros"]])
    ws.append(["(−) Abonos contables no en banco", bk["abonos_no_banco"]])
    ws.append(["(+) Cargos contables no en banco", bk["cargos_no_banco"]])
    ws.append(["= Saldo en bancos", bk["saldo_bancos"]])
    ws.append(["Diferencia en bancos", bk["diferencia"]])

    wc = wb.create_sheet("Conciliados")
    wc.append(["Fecha contable", "Concepto contable", "Documento", "Monto", "Fecha extracto", "Descripción extracto", "Cruce"])
    for c in d["conciliados"]:
        wc.append([c["fecha_c"], c["concepto_c"], c["documento"], c["monto"], c["fecha_e"], c["descripcion_e"], c["match_tipo"]])

    wp = wb.create_sheet("Por conciliar")
    wp.append(["EN LIBROS, NO EN BANCO"])
    wp.append(["Fecha", "Concepto", "Documento", "Monto"])
    for m in d["solo_libros"]:
        wp.append([m["fecha"], m["concepto"], m["documento"], m["monto"]])
    wp.append([])
    wp.append(["EN BANCO, NO EN LIBROS"])
    wp.append(["Fecha", "Descripción", "Código", "Monto"])
    for m in d["solo_banco"]:
        wp.append([m["fecha"], m["concepto"], m["codigo"], m["monto"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=conciliacion_{d['mes']}.xlsx"})
