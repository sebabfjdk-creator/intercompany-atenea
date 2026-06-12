"""Endpoints de datos para los tableros (lectura) y configuración (escritura)."""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.services import config_service, queries
from db.base import get_db
from db.models import ROLES, User

router = APIRouter(prefix="/api", tags=["data"])


def require_editor(user: User = Depends(get_current_user)) -> User:
    if user.rol not in ROLES:
        raise HTTPException(403, "No autorizado para editar la configuración")
    return user


class GrupoIn(BaseModel):
    grupo: str
    tipo: str
    tipo_relacion: str = "n_a_n"
    cuentas_co: list[str] = []
    cuentas_es: list[str] = []


class HomologacionIn(BaseModel):
    grupos: list[GrupoIn]


class ToleranciaIn(BaseModel):
    tolerancia_abs_cop: float
    tolerancia_pct: float


@router.get("/estado-datos")
def estado_datos(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.estado_datos(db)


def _xlsx_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/comparativa/export")
def export_comparativa(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    data = queries.comparativa(db)
    periodos = data["periodos"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparativa"
    header = ["Grupo", "Tipo"]
    for p in periodos:
        header += [f"{p} CO", f"{p} ES", f"{p} Dif", f"{p} Estado"]
    ws.append(header)
    for f in data["filas"]:
        row = [f["grupo"], f["tipo"]]
        for p in periodos:
            c = f["celdas"].get(p, {})
            row += [c.get("co", 0), c.get("es", 0), c.get("dif", 0), c.get("estado", "")]
        ws.append(row)
    return _xlsx_response(wb, "comparativa.xlsx")


@router.get("/comparativa/detalle-grupo")
def detalle_grupo(grupo: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.detalle_grupo(db, grupo)


@router.get("/comparativa/movimientos-cuenta")
def movimientos_cuenta(pais: str, cuenta: str, periodo: str | None = None, nit: str | None = None,
                       db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.movimientos_cuenta(db, pais, cuenta, periodo, nit)


@router.get("/comparativa")
def comparativa(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.comparativa(db)


@router.get("/resumen")
def resumen(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.resumen(db)


@router.get("/excepciones")
def excepciones(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.excepciones(db)


@router.get("/terceros")
def terceros(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.terceros(db)


@router.get("/config/homologacion")
def homologacion(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.homologacion(db)


@router.get("/auditoria")
def auditoria(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.auditoria(db)


@router.get("/anomalias")
def anomalias(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return queries.anomalias(db)


@router.get("/resumen/export")
def export_resumen(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    data = queries.resumen(db)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Resumen"
    ws.append(["Rubro", "Grupos", "CO", "ES", "Diferencia"])
    for r in data["rubros"]:
        ws.append([r["tipo"], r["grupos"], r["co"], r["es"], r["dif"]])
    return _xlsx_response(wb, "resumen.xlsx")


@router.get("/excepciones/export")
def export_excepciones(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    data = queries.excepciones(db)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Excepciones"
    ws.append(["Grupo", "Tipo", "Periodo", "CO", "ES", "Diferencia", "%", "Causa"])
    for e in data:
        ws.append([e["grupo"], e["tipo"], e["periodo"], e["total_co"], e["total_es"],
                   e["diferencia"], e["pct"], e["causa"] or ""])
    return _xlsx_response(wb, "excepciones.xlsx")


@router.get("/terceros/export")
def export_terceros(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    data = queries.terceros(db)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Terceros"
    ws.append(["Cuenta ES", "Nombre fiscal", "NIF norm.", "NIT Colombia", "Tipo"])
    for t in data["items"]:
        ws.append([t["cuenta_es"], t["nombre_fiscal"], t["nif_normalizado"], t["nit_colombia"], t["tipo"]])
    return _xlsx_response(wb, "terceros.xlsx")


# ---------- Configuración editable (homologación + tolerancias) ----------
@router.put("/config/homologacion")
def guardar_homologacion(body: HomologacionIn, db: Session = Depends(get_db), user: User = Depends(require_editor)):
    try:
        return config_service.save_homologacion(db, [g.model_dump() for g in body.grupos], user.id)
    except ValueError as e:
        raise HTTPException(422, str(e))


class MoverCuentaIn(BaseModel):
    cuenta: str
    pais: str               # CO | ES
    grupo_origen: str
    grupo_destino: str


@router.get("/config/conflictos")
def conflictos_homologacion(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    """Cuentas homologadas en más de un grupo (exacto o por wildcard) — doble conteo."""
    return config_service.cuentas_multiples_grupos(db)


@router.post("/config/homologacion/mover")
def mover_cuenta(body: MoverCuentaIn, db: Session = Depends(get_db), user: User = Depends(require_editor)):
    """Drag & drop: mueve una cuenta de un grupo a otro. Auto-guardado + auditoría."""
    try:
        return config_service.mover_cuenta(db, body.cuenta, body.pais,
                                           body.grupo_origen, body.grupo_destino, user.id)
    except ValueError as e:
        raise HTTPException(422, str(e))


@router.put("/config/tolerancia")
def guardar_tolerancia(body: ToleranciaIn, db: Session = Depends(get_db), user: User = Depends(require_editor)):
    config_service.set_tolerancia(db, body.tolerancia_abs_cop, body.tolerancia_pct, user.id)
    return {"ok": True, "tolerancia_abs_cop": body.tolerancia_abs_cop, "tolerancia_pct": body.tolerancia_pct}


@router.post("/config/recalcular")
def recalcular(_: User = Depends(require_editor)):
    # Arquitectura de cálculo en vivo: los tableros leen account_mapping en cada
    # request, así que el recálculo es automático. Endpoint provisto por compatibilidad.
    return {"ok": True, "modo": "live", "mensaje": "Los tableros se recalculan en vivo en cada consulta."}


@router.get("/config/homologacion/export")
def exportar_homologacion(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    data = config_service.get_homologacion(db)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Homologacion"
    ws.append(["Grupo", "Tipo", "Relacion", "Cuentas Colombia", "Cuentas Espana"])
    for g in data["grupos"]:
        ws.append([g["grupo"], g["tipo"], g["tipo_relacion"],
                   ", ".join(g["cuentas_co"]), ", ".join(g["cuentas_es"])])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=homologacion.xlsx"},
    )
