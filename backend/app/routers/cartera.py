"""Endpoints del módulo Cartera 360° (clientes España 430/431)."""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.services import cartera_service as svc
from db.base import get_db
from db.models import User

router = APIRouter(prefix="/api/cartera", tags=["cartera"])


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    return svc.dashboard(db)


@router.get("/export")
def export(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    import openpyxl
    d = svc.dashboard(db)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Resumen ejecutivo"
    ws.append(["Indicador", "Valor"])
    k = d["kpis"]
    for label, val in [
        ("Cartera total (COP)", k["cartera_total"]), ("Clientes con saldo", k["clientes"]),
        ("Dudoso cobro 431 (COP)", k["dudoso_431"]), ("Provisionales ES (COP)", k["provisional_es"]),
        ("Monto antigüedad >90d (COP)", k["monto_critico_90"]), ("Clientes en riesgo (>90d)", k["clientes_riesgo"]),
        ("Provisión recomendada (COP)", k["provision_recomendada"]), ("Concentración mayor cliente (%)", k["concentracion_top1"]),
    ]:
        ws.append([label, val])
    ws.append([])
    ws.append(["Análisis automático"])
    for linea in d["analisis"]:
        ws.append([linea])

    wa = wb.create_sheet("Aging (antiguedad)")
    wa.append(["Tramo", "Saldo (COP)", "Clientes"])
    for a in d["aging"]:
        wa.append([a["bucket"], a["saldo"], a["clientes"]])

    wc = wb.create_sheet("Matriz de cobrabilidad")
    wc.append(["NIT", "Cliente", "Saldo", "Saldo 430", "Saldo 431", "Días", "Antigüedad", "Riesgo"])
    for c in d["clientes"]:
        wc.append([c["nit"], c["nombre"], c["saldo"], c["saldo_430"], c["saldo_431"],
                   c["dias"], c["antiguedad"], c["riesgo"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cartera_360.xlsx"})
