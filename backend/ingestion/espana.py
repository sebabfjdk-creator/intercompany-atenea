"""Adapter de ingesta — España (DELSOL), formato 'Libro Mayor' impreso.

Estructura observada en los datos reales (hoja AteneaEneroMvto / AteneaFebrero-MarzoMvti):

    Cuenta: [100 ...]                          <- marcador de sección (se ignora)
    100.0.0.000 CAPITAL ...  | Anterior: | ... <- apertura de cuenta + saldo anterior
    2026-01-01 | 1 | 1 | 1 | ASIENTO ...| | Debe | Haber | Saldo   <- movimiento
    ...
    | Total:           | ...                   <- se ignora
    | Total de cuenta: | debe | haber | saldo   <- usado para validación

Columnas: Fecha | Dia | Asto. | Ord | Concepto | Docum. | Debe | Haber | Saldo

Notas:
- Los importes en este archivo "Ok consolidado" son numéricos nativos, pero
  `parse_es_number` se aplica igual para tolerar futuros export crudos en formato ES.
- El código de cuenta se propaga a cada movimiento del bloque.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from ingestion.utils import es_codigo_cuenta_es, parse_es_number

# Markers que aparecen en la columna 'Docum.' y no son movimientos.
#   'Anterior:'        -> saldo anterior (rolling) de la cuenta, en la fila del código
#   'Total:'           -> total de ESA subcuenta (código completo). Incluye el anterior.
#   'Total de cuenta:' -> total del grupo padre (agrega subcuentas). Se ignora.
_MARKERS_RESUMEN = {"anterior:", "total:", "total de cuenta:"}


@dataclass
class MovimientoES:
    cuenta_es: str
    cuenta_nombre: str
    fecha: datetime | None
    concepto: str
    docum: str
    debe: float
    haber: float
    saldo: float
    neto: float = field(init=False)

    def __post_init__(self):
        self.neto = round(self.debe - self.haber, 2)


@dataclass
class CuentaES:
    codigo: str
    nombre: str
    ant_debe: float = 0.0     # saldo anterior (rolling), fila 'Anterior:'
    ant_haber: float = 0.0
    ant_saldo: float = 0.0
    total_debe: float = 0.0   # fila 'Total:' de esta subcuenta (incluye anterior)
    total_haber: float = 0.0
    saldo_final: float = 0.0
    movimientos: list[MovimientoES] = field(default_factory=list)


def _find_header_row(ws, max_scan: int = 30) -> tuple[int, dict[str, int]]:
    """Localiza la fila de cabecera 'Fecha|Dia|...|Debe|Haber|Saldo' y devuelve
    (índice_fila_1based, mapa nombre->índice_columna_0based)."""
    wanted = {"fecha", "concepto", "debe", "haber", "saldo", "docum.", "dia"}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        labels = {str(c).strip().lower() for c in row if c is not None}
        if {"fecha", "debe", "haber", "saldo"}.issubset(labels):
            colmap = {}
            for i, c in enumerate(row):
                if c is not None:
                    colmap[str(c).strip().lower()] = i
            return r, colmap
    raise ValueError("No se encontró la fila de cabecera (Fecha/Debe/Haber/Saldo)")


def parse_espana_movimientos(path, sheet: str) -> list[CuentaES]:
    """Parsea una hoja de movimientos DELSOL y devuelve la lista de cuentas
    con sus movimientos y totales de control."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        header_row, colmap = _find_header_row(ws)

        c_fecha = colmap.get("fecha", 0)
        c_concepto = colmap.get("concepto", 4)
        c_docum = colmap.get("docum.", colmap.get("docum", 5))
        c_debe = colmap.get("debe", 6)
        c_haber = colmap.get("haber", 7)
        c_saldo = colmap.get("saldo", 8)

        cuentas: list[CuentaES] = []
        actual: CuentaES | None = None

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue
            col0 = row[c_fecha]
            docum = (str(row[c_docum]).strip() if c_docum < len(row) and row[c_docum] is not None else "")

            debe_cell = parse_es_number(row[c_debe] if c_debe < len(row) else 0)
            haber_cell = parse_es_number(row[c_haber] if c_haber < len(row) else 0)
            saldo_cell = parse_es_number(row[c_saldo] if c_saldo < len(row) else 0)

            # ¿apertura de cuenta? (col0 == 'NNN.N.N.NNN NOMBRE ...')
            cod = es_codigo_cuenta_es(col0)
            if cod:
                nombre = str(col0).strip()[len(cod):].strip()
                actual = CuentaES(codigo=cod, nombre=nombre)
                cuentas.append(actual)
                # esa misma fila trae 'Anterior:' con el saldo anterior (rolling)
                if docum.lower() == "anterior:":
                    actual.ant_debe = debe_cell
                    actual.ant_haber = haber_cell
                    actual.ant_saldo = saldo_cell
                continue

            marker = docum.lower()
            if marker in _MARKERS_RESUMEN:
                # 'Total:' = total de esta subcuenta (control). 'Total de cuenta:' = grupo padre (ignorar).
                if marker == "total:" and actual is not None:
                    actual.total_debe = debe_cell
                    actual.total_haber = haber_cell
                    actual.saldo_final = saldo_cell
                continue

            # ¿marcador de sección 'Cuenta: [...]' u otra metadata sin cuenta activa?
            if actual is None:
                continue

            # ¿es un movimiento? requiere fecha en col0
            fecha = col0 if isinstance(col0, datetime) else None
            if fecha is None and debe_cell == 0 and haber_cell == 0:
                continue  # fila vacía / metadata intermedia

            mov = MovimientoES(
                cuenta_es=actual.codigo,
                cuenta_nombre=actual.nombre,
                fecha=fecha,
                concepto=str(row[c_concepto]).strip() if c_concepto < len(row) and row[c_concepto] is not None else "",
                docum=docum,
                debe=debe_cell,
                haber=haber_cell,
                saldo=saldo_cell,
            )
            actual.movimientos.append(mov)

        return cuentas
    finally:
        wb.close()


def validar_totales(cuentas: list[CuentaES], tol: float = 0.5) -> list[dict]:
    """Valida la integridad del parseo por cuenta usando la identidad del Libro
    Mayor:  Total: = saldo_anterior + Σ movimientos.

    Devuelve la lista de descuadres (vacía => todo cuadra). `tol` en COP absorbe
    redondeos del export.
    """
    descuadres = []
    for c in cuentas:
        exp_debe = round(c.ant_debe + sum(m.debe for m in c.movimientos), 2)
        exp_haber = round(c.ant_haber + sum(m.haber for m in c.movimientos), 2)
        if abs(exp_debe - c.total_debe) > tol or abs(exp_haber - c.total_haber) > tol:
            descuadres.append({
                "cuenta": c.codigo,
                "exp_debe": exp_debe, "total_debe": c.total_debe,
                "exp_haber": exp_haber, "total_haber": c.total_haber,
            })
    return descuadres
