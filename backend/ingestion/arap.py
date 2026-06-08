"""Parsers del módulo AR/AP (Cuentas por Cobrar y Pagar).

Archivo fuente: CarteraYPasivos.xlsx con 4 hojas:
  - CarteraAtenea (España AR, cuenta 430)   | CXPAtenea  (España AP, cuenta 410)  -> formato DELSOL
  - CarteraNeuron (Colombia AR, 1305/2805)  | CXPNeuron  (Colombia AP, 22xx)      -> formato Siesa

España (DELSOL): bloques por subcuenta NNN.N.N.NNN. El saldo del tercero es la
columna Saldo de la fila 'Total:' que cierra cada subcuenta (no el acumulado).
Las subcuentas con relleno AMARILLO (FFFFFF00) son provisionales (430.9.x / 431.9.9.x,
"FACTURAS PEND. EMITIR") -> es_provisional=True, no cruzan.

Colombia (Siesa): el saldo por tercero es la fila resumen de cada NIT (sin Referencia;
en el Excel va en fuente azul FF000080). saldo_neto = Σ(1305) + Σ(2805) por NIT.
Saldo negativo en 1305 -> error_contabilizacion (debería estar en 2805).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime

import openpyxl

from ingestion.utils import es_codigo_cuenta_es, parse_es_number

# Mapeo de prefijo de documento -> tipo legible
_TIPO_DOC = {
    "FA": "Factura", "FE": "Factura", "FV": "Factura",
    "NC": "Nota crédito", "NCA": "Nota crédito", "ND": "Nota débito",
    "RC": "Recibo/Pago", "CL": "Pago", "PR": "Provisión",
}


def normalizar_doc(doc) -> str:
    """Número de documento sin guiones ni espacios, en mayúsculas (para cruce CO<->ES)."""
    return re.sub(r"[\s\-]", "", str(doc or "")).upper()


def tipo_documento(doc) -> str:
    s = normalizar_doc(doc)
    m = re.match(r"^([A-Z]+)", s)
    if not m:
        return "Otro" if s else ""
    return _TIPO_DOC.get(m.group(1), "Otro")

# --- España: índices de columna fijos (sin fila de cabecera) ---
_ES_DOCUM, _ES_DEBE, _ES_HABER, _ES_SALDO = 5, 6, 7, 8
_AMARILLO = "FFFFFF00"


@dataclass
class MovArAp:
    cuenta: str
    fecha: object  # datetime | None
    concepto: str
    debe: float
    haber: float
    saldo: float
    documento: str = ""


@dataclass
class SaldoTerceroES:
    cuenta_es: str
    nombre: str
    saldo: float
    es_provisional: bool
    movimientos: list = None  # list[MovArAp] del bloque (detalle)


@dataclass
class SaldoTerceroCO:
    nit: str
    nombre: str
    saldo_1305: float
    saldo_2805: float
    saldo_22xx: float

    @property
    def saldo_neto(self) -> float:
        return round(self.saldo_1305 + self.saldo_2805 + self.saldo_22xx, 2)

    @property
    def error_contabilizacion(self) -> bool:
        return self.saldo_1305 < -0.5  # negativo en 1305 = posible error


def _cell_amarilla(cell) -> bool:
    f = getattr(cell, "fill", None)
    if not f or f.patternType != "solid":
        return False
    rgb = getattr(f.fgColor, "rgb", None)
    return str(rgb) == _AMARILLO


def parse_arap_espana(path, sheet: str) -> list[SaldoTerceroES]:
    """Saldo final por subcuenta (tercero) de una hoja DELSOL AR/AP."""
    wb = openpyxl.load_workbook(path, data_only=True)  # sin read_only -> acceso a fills
    try:
        ws = wb[sheet]
        out: list[SaldoTerceroES] = []
        actual: SaldoTerceroES | None = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9):
            c0 = row[0]
            cod = es_codigo_cuenta_es(c0.value)
            if cod:
                nombre = str(c0.value).strip()[len(cod):].strip()
                actual = SaldoTerceroES(cuenta_es=cod, nombre=nombre, saldo=0.0,
                                        es_provisional=_cell_amarilla(c0), movimientos=[])
                out.append(actual)
                continue
            docum = str(row[_ES_DOCUM].value).strip().lower() if row[_ES_DOCUM].value is not None else ""
            if docum == "total:" and actual is not None:
                actual.saldo = parse_es_number(row[_ES_SALDO].value)
                continue
            # fila de movimiento: fecha en col0
            if actual is not None and isinstance(c0.value, datetime):
                actual.movimientos.append(MovArAp(
                    cuenta=actual.cuenta_es, fecha=c0.value,
                    concepto=str(row[4].value).strip() if row[4].value is not None else "",
                    documento=str(row[_ES_DOCUM].value).strip() if row[_ES_DOCUM].value is not None else "",
                    debe=parse_es_number(row[_ES_DEBE].value),
                    haber=parse_es_number(row[_ES_HABER].value),
                    saldo=parse_es_number(row[_ES_SALDO].value),
                ))
        return out
    finally:
        wb.close()


def parse_arap_colombia(path, sheet: str) -> list[SaldoTerceroCO]:
    """Saldo por NIT (filas resumen sin referencia), agregando 1305/2805/22xx."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        # cabecera en fila 1: Cuenta|Nombre cuenta|Tercero|Nombre tercero|Referencia|...|Saldo
        c_cta, c_terc, c_nterc, c_ref, c_saldo = 0, 2, 3, 4, 10
        agg: dict[str, SaldoTerceroCO] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            nit = row[c_terc] if c_terc < len(row) else None
            if nit is None or str(nit).strip() == "":
                continue
            ref = row[c_ref] if c_ref < len(row) else None
            if ref is not None and str(ref).strip() != "":
                continue  # fila de detalle (gris); el saldo está en la fila resumen (azul, sin ref)
            cuenta = str(row[c_cta]).strip() if row[c_cta] is not None else ""
            saldo = parse_es_number(row[c_saldo] if c_saldo < len(row) else 0)
            nit_s = str(nit).strip()
            t = agg.get(nit_s)
            if t is None:
                t = SaldoTerceroCO(nit=nit_s, nombre=str(row[c_nterc]).strip() if row[c_nterc] else "",
                                   saldo_1305=0.0, saldo_2805=0.0, saldo_22xx=0.0)
                agg[nit_s] = t
            if cuenta.startswith("1305"):
                t.saldo_1305 = round(t.saldo_1305 + saldo, 2)
            elif cuenta.startswith("2805"):
                t.saldo_2805 = round(t.saldo_2805 + saldo, 2)
            elif cuenta.startswith("22"):
                t.saldo_22xx = round(t.saldo_22xx + saldo, 2)
        return list(agg.values())
    finally:
        wb.close()


def parse_arap_colombia_movimientos(path, sheet: str) -> list[tuple[str, MovArAp]]:
    """Devuelve (nit, MovArAp) por cada fila de detalle (con Referencia)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        c_cta, c_ncta, c_terc, c_ref, c_fec, c_deb, c_cred, c_saldo = 0, 1, 2, 4, 5, 8, 9, 10
        out: list[tuple[str, MovArAp]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            nit = row[c_terc] if c_terc < len(row) else None
            ref = row[c_ref] if c_ref < len(row) else None
            if nit is None or str(nit).strip() == "":
                continue
            if ref is None or str(ref).strip() == "":
                continue  # solo filas de detalle (con referencia)
            fec = row[c_fec] if c_fec < len(row) else None
            out.append((str(nit).strip(), MovArAp(
                cuenta=str(row[c_cta]).strip() if row[c_cta] is not None else "",
                fecha=fec if isinstance(fec, datetime) else None,
                concepto=(str(row[c_ncta]).strip() if row[c_ncta] else ""),
                documento=str(ref).strip() if ref else "",
                debe=parse_es_number(row[c_deb] if c_deb < len(row) else 0),
                haber=parse_es_number(row[c_cred] if c_cred < len(row) else 0),
                saldo=parse_es_number(row[c_saldo] if c_saldo < len(row) else 0),
            )))
        return out
    finally:
        wb.close()
