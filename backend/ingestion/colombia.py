"""Adapter de ingesta — Colombia (Siesa).

Dos hojas por periodo:

* Balance_<periodo>: cuentas jerárquicas (nivel por longitud de código:
  1 -> 11 -> 1120 -> 112005 -> 11200501). Columnas tras 2 filas de cabecera:
  Cód. | Nombre |  | Saldo Ant. | Débitos | Créditos | Saldo actual.
  Identidad de control:  Saldo actual = Saldo Ant + Débitos - Créditos.
  Es la fuente para el cruce PYG por cuenta/mes.

* Mvto_<periodo>: detalle por tercero (NIT). Columnas:
  Cuenta | Nombre cuenta | Tercero | Nombre tercero | Referencia |
  F. creación | F. pago | Saldo anterior | Débito | Crédito | Saldo.
  Por cada tercero hay una fila-resumen (sin Referencia) y filas-detalle
  (con Referencia). Para AR/AP usamos las filas-detalle. Fuente para terceros.

Números: formato anglosajón nativo (punto decimal). `parse_es_number` los
admite igual.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime

import openpyxl

from ingestion.utils import parse_es_number


@dataclass
class CuentaBalanceCO:
    codigo: str
    nombre: str
    nivel: int          # longitud del código => nivel jerárquico
    saldo_ant: float
    debitos: float
    creditos: float
    saldo_actual: float

    @property
    def neto(self) -> float:
        """Movimiento del periodo (débitos - créditos)."""
        return round(self.debitos - self.creditos, 2)


@dataclass
class MovimientoCO:
    cuenta: str
    nombre_cuenta: str
    nit: str
    nombre_tercero: str
    referencia: str
    fecha: datetime | None
    saldo_anterior: float
    debito: float
    credito: float
    saldo: float
    neto: float = field(init=False)

    def __post_init__(self):
        self.neto = round(self.debito - self.credito, 2)


def _find_header(ws, needles: set[str], max_scan: int = 6):
    """Devuelve (fila_1based, {label_lower: idx0}) de la primera fila que
    contiene todos los `needles`."""
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        labels = {str(c).strip().lower(): i for i, c in enumerate(row) if c is not None}
        if needles.issubset(set(labels)):
            return r, labels
    raise ValueError(f"No se encontró cabecera con {needles}")


def parse_colombia_balance(path, sheet: str) -> list[CuentaBalanceCO]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        # cabecera real en la 2ª fila: 'cód.'/'nombre'/'saldo ant.'/'débitos'...
        header_row, labels = _find_header(
            ws, {"débitos", "créditos", "saldo actual"}, max_scan=6
        )
        c_cod = labels.get("cód.", labels.get("cod.", 0))
        c_nom = labels.get("nombre", 1)
        c_ant = labels.get("saldo ant.", 3)
        c_deb = labels["débitos"]
        c_cred = labels["créditos"]
        c_act = labels["saldo actual"]

        cuentas: list[CuentaBalanceCO] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cod = row[c_cod] if c_cod < len(row) else None
            if cod is None or str(cod).strip() == "":
                continue
            cod = str(cod).strip()
            if not cod.isdigit():
                continue
            cuentas.append(CuentaBalanceCO(
                codigo=cod,
                nombre=str(row[c_nom]).strip() if c_nom < len(row) and row[c_nom] is not None else "",
                nivel=len(cod),
                saldo_ant=parse_es_number(row[c_ant] if c_ant < len(row) else 0),
                debitos=parse_es_number(row[c_deb] if c_deb < len(row) else 0),
                creditos=parse_es_number(row[c_cred] if c_cred < len(row) else 0),
                saldo_actual=parse_es_number(row[c_act] if c_act < len(row) else 0),
            ))
        return cuentas
    finally:
        wb.close()


def parse_colombia_movimientos(path, sheet: str, solo_detalle: bool = True) -> list[MovimientoCO]:
    """Extrae movimientos por tercero. Con `solo_detalle=True` devuelve únicamente
    las filas con Referencia (evita duplicar con las filas-resumen por tercero)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        header_row, labels = _find_header(
            ws, {"cuenta", "tercero", "débito", "crédito"}, max_scan=3
        )
        c_cta = labels["cuenta"]
        c_ncta = labels.get("nombre cuenta", 1)
        c_terc = labels["tercero"]
        c_nterc = labels.get("nombre tercero", 3)
        c_ref = labels.get("referencia", 4)
        c_fcre = labels.get("f. creación", 5)
        c_ant = labels.get("saldo anterior", 7)
        c_deb = labels["débito"]
        c_cred = labels["crédito"]
        c_sal = labels.get("saldo", 10)

        movs: list[MovimientoCO] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            terc = row[c_terc] if c_terc < len(row) else None
            if terc is None or str(terc).strip() == "":
                continue
            ref = row[c_ref] if c_ref < len(row) else None
            if solo_detalle and (ref is None or str(ref).strip() == ""):
                continue
            fecha = row[c_fcre] if c_fcre < len(row) else None
            movs.append(MovimientoCO(
                cuenta=str(row[c_cta]).strip() if c_cta < len(row) and row[c_cta] is not None else "",
                nombre_cuenta=str(row[c_ncta]).strip() if c_ncta < len(row) and row[c_ncta] is not None else "",
                nit=str(terc).strip(),
                nombre_tercero=str(row[c_nterc]).strip() if c_nterc < len(row) and row[c_nterc] is not None else "",
                referencia=str(ref).strip() if ref is not None else "",
                fecha=fecha if isinstance(fecha, datetime) else None,
                saldo_anterior=parse_es_number(row[c_ant] if c_ant < len(row) else 0),
                debito=parse_es_number(row[c_deb] if c_deb < len(row) else 0),
                credito=parse_es_number(row[c_cred] if c_cred < len(row) else 0),
                saldo=parse_es_number(row[c_sal] if c_sal < len(row) else 0),
            ))
        return movs
    finally:
        wb.close()


def validar_balance(cuentas: list[CuentaBalanceCO], tol: float = 1.0) -> list[dict]:
    """Verifica la identidad Saldo actual = Saldo Ant + Débitos - Créditos."""
    descuadres = []
    for c in cuentas:
        esperado = round(c.saldo_ant + c.debitos - c.creditos, 2)
        if abs(esperado - c.saldo_actual) > tol:
            descuadres.append({
                "cuenta": c.codigo, "esperado": esperado, "saldo_actual": c.saldo_actual,
            })
    return descuadres
