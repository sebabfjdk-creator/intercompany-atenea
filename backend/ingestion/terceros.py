"""Adapter de ingesta — Puente de terceros (NIF España <-> NIT Colombia).

Fuentes:
* Homologacion.xlsx, hoja 'Puente Terceros' (cabecera en fila 4): fuente de verdad,
  ya trae 'NIF normalizado' y 'NIT Colombia' precalculados.
* Reporte_de_terceros.xlsx, hojas 'Clientes' y 'proveedor': catálogo DELSOL crudo
  (CÓDIGO 430/431/410, NOMBRE FISCAL, N.I.F. = NIT Colombia).

`normalizar_nif()` se valida contra la columna precalculada del puente.
"""
from __future__ import annotations

import unicodedata
from dataclasses import dataclass

import openpyxl

from ingestion.utils import normalizar_nif


def _norm_label(s) -> str:
    """Normaliza una etiqueta de cabecera: minúsculas, sin acentos, sin espacios
    de borde. Robusto frente a NFC/NFD entre archivos."""
    s = str(s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


@dataclass
class TerceroPuente:
    tipo: str                  # Cliente | Proveedor
    cuenta_es: str
    nombre_fiscal: str
    nombre_comercial: str
    nif_original: str
    nif_normalizado: str       # precalculado en el archivo
    tipo_nif: str
    nit_colombia: str
    nombre_colombia: str


def _find_header(ws, needles: set[str], max_scan: int = 8):
    """Localiza la cabecera de forma insensible a acentos. Las claves devueltas
    están ya normalizadas (sin acentos) para indexar columnas."""
    needles = {_norm_label(n) for n in needles}
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        labels = {_norm_label(c): i for i, c in enumerate(row) if c is not None}
        if needles.issubset(set(labels)):
            return r, labels
    raise ValueError(f"No se encontró cabecera con {needles}")


def load_puente_terceros(path, sheet: str = "Puente Terceros") -> list[TerceroPuente]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        header_row, lb = _find_header(
            ws, {"cuenta espana", "nif normalizado", "nit colombia"}, max_scan=8
        )

        def col(name, default=None):
            return lb.get(name, default)

        c_tipo = col("tipo", 0)
        c_cta = col("cuenta espana")
        c_nf = col("nombre fiscal")
        c_nc = col("nombre comercial")
        c_nifo = col("n.i.f. original", col("nif original"))
        c_nifn = col("nif normalizado")
        c_tnif = col("tipo nif")
        c_nit = col("nit colombia")
        c_ncol = col("nombre colombia")

        out: list[TerceroPuente] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cta = row[c_cta] if c_cta is not None and c_cta < len(row) else None
            if cta is None or str(cta).strip() == "":
                continue

            def g(idx):
                return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""

            out.append(TerceroPuente(
                tipo=g(c_tipo),
                cuenta_es=g(c_cta),
                nombre_fiscal=g(c_nf),
                nombre_comercial=g(c_nc),
                nif_original=g(c_nifo),
                nif_normalizado=g(c_nifn),
                tipo_nif=g(c_tnif),
                nit_colombia=g(c_nit),
                nombre_colombia=g(c_ncol),
            ))
        return out
    finally:
        wb.close()


def load_reporte_terceros(path) -> list[dict]:
    """Catálogo DELSOL crudo: hojas 'Clientes' y 'proveedor'. Devuelve dicts con
    cuenta_es, nombre_fiscal, nif_original, nif_normalizado, tipo."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out: list[dict] = []
        for sheet, tipo in [("Clientes", "Cliente"), ("proveedor", "Proveedor")]:
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            # 'Clientes' usa CÓDIGO; 'proveedor' usa CUENTA para la cuenta ES
            header_row, lb = _find_header(ws, {"n.i.f.", "nombre fiscal"}, max_scan=4)
            c_cod = lb.get("codigo", lb.get("cuenta"))
            c_nf = lb.get("nombre fiscal")
            c_nif = lb["n.i.f."]
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                cod = row[c_cod] if c_cod < len(row) else None
                if cod is None or str(cod).strip() == "":
                    continue
                nif_o = str(row[c_nif]).strip() if c_nif < len(row) and row[c_nif] is not None else ""
                out.append({
                    "cuenta_es": str(cod).strip(),
                    "nombre_fiscal": str(row[c_nf]).strip() if c_nf is not None and c_nf < len(row) and row[c_nf] is not None else "",
                    "nif_original": nif_o,
                    "nif_normalizado": normalizar_nif(nif_o),
                    "tipo": tipo,
                })
        return out
    finally:
        wb.close()


def validar_normalizacion(puente: list[TerceroPuente]) -> list[dict]:
    """Compara nuestra normalizar_nif(nif_original) con la columna precalculada
    'NIF normalizado' del archivo. Devuelve discrepancias."""
    diffs = []
    for t in puente:
        if not t.nif_original:
            continue
        calc = normalizar_nif(t.nif_original)
        ref = t.nif_normalizado.upper().replace(" ", "")
        if ref and calc != ref:
            diffs.append({"nif": t.nif_original, "calc": calc, "ref": ref})
    return diffs
