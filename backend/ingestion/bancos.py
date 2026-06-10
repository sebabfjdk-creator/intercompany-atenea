"""Parsers de los dos insumos de la Conciliación Bancaria.

- Contable (BancosAtenea, formato Libro Mayor): col0 fecha, col4 concepto,
  col5 documento, col6 DÉBITO (ingresos), col7 CRÉDITO (pagos). Monto firmado =
  débito − crédito (+ ingreso, − pago).
- Extracto (BancoFebrero, tabla plana): col0 nº cuenta, col3 fecha YYYYMMDD,
  col5 valor firmado (− pago, + ingreso), col6 código, col7 descripción.

Cruzan por igualdad de monto firmado (un pago es negativo en ambos lados).
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime

import openpyxl

from ingestion.utils import parse_es_number


@dataclass
class MovBanco:
    fecha: datetime | None
    monto: float            # firmado: + ingreso, − pago
    concepto: str
    documento: str = ""
    codigo: str = ""
    numero_cuenta: str = ""


def _ws(path, sheet_hint: str | None):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    names = wb.sheetnames
    name = next((s for s in names if sheet_hint and sheet_hint.lower() in s.lower()), names[0])
    return wb, wb[name]


def parse_bancos_contable(path, sheet: str | None = "Bancos") -> list[MovBanco]:
    wb, ws = _ws(path, sheet)
    try:
        out: list[MovBanco] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            c0 = row[0] if row else None
            if not isinstance(c0, datetime):
                continue  # descarta cabeceras 'Cuenta: [...]' y filas vacías
            deb = parse_es_number(row[6] if len(row) > 6 else 0)
            cred = parse_es_number(row[7] if len(row) > 7 else 0)
            out.append(MovBanco(
                fecha=c0, monto=round(deb - cred, 2),
                concepto=str(row[4] or "")[:300] if len(row) > 4 else "",
                documento=str(row[5] or "")[:60] if len(row) > 5 else "",
            ))
        return out
    finally:
        wb.close()


def parse_banco_extracto(path, sheet: str | None = "Banco") -> list[MovBanco]:
    wb, ws = _ws(path, sheet)
    try:
        out: list[MovBanco] = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            f = str(row[3] or "").strip() if len(row) > 3 else ""
            if len(f) != 8 or not f.isdigit():
                continue
            out.append(MovBanco(
                fecha=datetime.strptime(f, "%Y%m%d"),
                monto=round(parse_es_number(row[5] if len(row) > 5 else 0), 2),
                concepto=str(row[7] or "")[:300] if len(row) > 7 else "",
                codigo=str(row[6] or "")[:40] if len(row) > 6 else "",
                numero_cuenta=str(row[0] or "")[:40] if row else "",
            ))
        return out
    finally:
        wb.close()


def mes_dominante(movs: list[MovBanco]) -> str | None:
    from collections import Counter
    c = Counter(m.fecha.strftime("%Y-%m") for m in movs if m.fecha)
    return c.most_common(1)[0][0] if c else None


def conciliar(contables: list[MovBanco], extracto: list[MovBanco], tol_dias: int = 3):
    """Empareja 1:1: (1) misma fecha y mismo monto firmado; (2) mismo monto y
    fecha dentro de ±tol_dias. Devuelve (pares, solo_libros_idx, solo_banco_idx)
    donde pares = [(i_cont, j_ext, 'exacto'|'fecha_dif')]."""
    ext_used = [False] * len(extracto)
    cont_done = [False] * len(contables)
    pares: list[tuple[int, int, str]] = []

    def fday(m):
        return m.fecha.date() if m.fecha else None

    for i, c in enumerate(contables):
        for j, e in enumerate(extracto):
            if not ext_used[j] and e.monto == c.monto and fday(e) == fday(c):
                pares.append((i, j, "exacto")); ext_used[j] = True; cont_done[i] = True
                break
    for i, c in enumerate(contables):
        if cont_done[i]:
            continue
        for j, e in enumerate(extracto):
            if ext_used[j] or e.monto != c.monto:
                continue
            dc, de = fday(c), fday(e)
            if dc and de and abs((de - dc).days) <= tol_dias:
                pares.append((i, j, "fecha_dif")); ext_used[j] = True; cont_done[i] = True
                break
    solo_libros = [i for i in range(len(contables)) if not cont_done[i]]
    solo_banco = [j for j in range(len(extracto)) if not ext_used[j]]
    return pares, solo_libros, solo_banco
