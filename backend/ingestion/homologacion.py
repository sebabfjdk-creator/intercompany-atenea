"""Adapter de ingesta — tabla de homologación de cuentas (CO <-> ES).

Hoja 'Gastos' (clase 5 CO <-> clase 6 ES): bloques separados por filas en blanco.
Cada bloque agrupa N cuentas Colombia y N cuentas España (relación N:N). La
columna País viene en mayúsc/minúsc variable ('Colombia'/'colombia'/'España').

Hoja 'Ingresos' (clase 4 CO <-> clase 7 ES): un bloque de cuentas CO seguido de
un bloque de cuentas ES (no son pares fila-a-fila). Se modela como un grupo de
ingresos con todas las cuentas CO vs todas las ES. La segmentación fina por
sub-rubro (la que insinúa el bosquejo) queda pendiente de aclaración de datos.

Devuelve `GrupoHomologado` con listas de códigos CO y ES, listo para alimentar
`account_mapping` y el motor de conciliación.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

from ingestion.utils import ACCOUNT_ES_RE

_ES_CODE_INLINE = re.compile(r"^\d{3}\.\d\.\d\.\d{3}")


def _clean_code(v) -> str:
    return str(v).strip() if v is not None else ""


@dataclass
class GrupoHomologado:
    grupo: str                       # etiqueta legible
    tipo: str                        # 'gasto' | 'ingreso'
    cuentas_co: list[str] = field(default_factory=list)
    cuentas_es: list[str] = field(default_factory=list)
    descripcion: str = ""

    @property
    def tipo_relacion(self) -> str:
        nco, nes = len(self.cuentas_co), len(self.cuentas_es)
        if nco == 1 and nes == 1:
            return "directa"
        if nco == 0 or nes == 0:
            return "sin_par"
        return "n_a_n"


def _load_gastos(ws) -> list[GrupoHomologado]:
    grupos: list[GrupoHomologado] = []
    cur_co: list[tuple[str, str]] = []
    cur_es: list[str] = []

    def flush():
        nonlocal cur_co, cur_es
        if cur_co or cur_es:
            label = cur_co[0][1] if cur_co else (cur_es[0] if cur_es else "grupo")
            grupos.append(GrupoHomologado(
                grupo=label or (cur_co[0][0] if cur_co else "grupo"),
                tipo="gasto",
                cuentas_co=[c for c, _ in cur_co],
                cuentas_es=list(cur_es),
                descripcion=label,
            ))
        cur_co, cur_es = [], []

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        pais = str(row[0]).strip().lower() if row[0] is not None else ""
        codigo = _clean_code(row[1])
        desc = _clean_code(row[2])
        if not pais and not codigo:
            flush()
            continue
        if pais.startswith("colombia"):
            if codigo:
                cur_co.append((codigo, desc))
        elif pais.startswith("espa"):
            if codigo:
                cur_es.append(codigo)
    flush()
    return [g for g in grupos if g.cuentas_co or g.cuentas_es]


def _load_ingresos(ws) -> list[GrupoHomologado]:
    """La hoja Ingresos lista cuentas CO (clase 4) y ES (clase 7) en bloques.
    Se agrupan en un único rubro 'Ingresos operacionales'."""
    co, es = [], []
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        code = _clean_code(row[0])
        if not code:
            continue
        if _ES_CODE_INLINE.match(code) or ACCOUNT_ES_RE.match(code):
            es.append(code)
        elif code[0].isdigit() and code[0] == "4":
            co.append(code)
    if not co and not es:
        return []
    return [GrupoHomologado(
        grupo="Ingresos operacionales",
        tipo="ingreso",
        cuentas_co=co,
        cuentas_es=es,
        descripcion="Ingresos (clase 4 CO <-> clase 7 ES)",
    )]


def load_homologacion(path) -> list[GrupoHomologado]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        grupos: list[GrupoHomologado] = []
        if "Gastos" in wb.sheetnames:
            grupos += _load_gastos(wb["Gastos"])
        if "Ingresos" in wb.sheetnames:
            grupos += _load_ingresos(wb["Ingresos"])
        return grupos
    finally:
        wb.close()
